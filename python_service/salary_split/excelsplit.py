#!/usr/bin/env python
# coding: utf-8

# In[1]:


import openpyxl
import re
from copy import copy
from openpyxl.utils import get_column_letter, column_index_from_string

import requests
import os

def upload_and_send_document(file_path, send_to_employee_id, document_parent_id = None):
    # ca_bundle_path = "/etc/ssl/cert.pem"

    # Step 1: Send POST request to get upload URL and document ID
    auth_token_1 = "4a015c8780bd41e4bb621213ac032058"
    url_1 = "https://working-day.online:8080/v1/documents/upload"
    headers_1 = {"Authorization": "Bearer " + auth_token_1}
    
    response_1 = requests.post(url_1, headers=headers_1, verify=False)
    response_1.raise_for_status()  # Raise an exception for HTTP errors
    response_json_1 = response_1.json()
    
    upload_url = response_json_1["url"]
    document_id = response_json_1["id"]
    
    # Step 2: Upload the file
    with open(file_path, 'rb') as file_data:
        response_2 = requests.put(upload_url, data=file_data, verify=False)
        response_2.raise_for_status()  # Raise an exception for HTTP errors
    
    # Step 3: Send POST request to send the document
    auth_token_2 = "4a015c8780bd41e4bb621213ac032058"
    url_3 = "https://working-day.online:8080/v1/documents/send"
    headers_3 = {
        "Authorization": "Bearer " + auth_token_2,
        "Content-Type": "application/json"
    }
    payload_3 = {
        "employee_ids": [send_to_employee_id],
        "document": {"id": document_id, "name": "Расчетный лист"}
    }
    if document_parent_id:
        payload_3["document"]["parent_id"] = document_parent_id
    
    response_3 = requests.post(url_3, json=payload_3, headers=headers_3, verify=False)
    response_3.raise_for_status()  # Raise an exception for HTTP errors
    
    # Step 4: Return the document ID
    return document_id



# In[2]:


def normal_width(filename):
    #функция нормализует ширину таблицы
    wb = openpyxl.load_workbook(filename)

    # Выбор активного листа (первого листа в книге)
    ws = wb.active

    # Установка заданной ширины для всех столбцов (например, 10)
    desired_width = 3

    # Обработка каждого столбца
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_length = 0

        # Находим максимальную длину содержимого в столбце, включая объединенные ячейки
        for cell in column_cells:
            try:
                if isinstance(cell, openpyxl.cell.MergedCell):
                    value = cell.value  # Используем значение из объединенной ячейки
                else:
                    value = cell.value

                if isinstance(value, str):
                    max_length = max(max_length, len(value))
                elif isinstance(value, (int, float)):
                    max_length = max(max_length, len(str(value)))
            except TypeError:
                pass

        # Устанавливаем заданную ширину для столбца
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = desired_width

    # Сохраняем изменения в файле
    wb.save(filename)


def split_xlsx(file_path, output_path):
    #разбивает общий эксель с расчетными листами на отдельные под каждого сотрудника 
    #название файла для сотрудника - фио+id 1C, формат - xlsx
    #возвращает список созданных файлов
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    desired_width = 0.5
    for col in ws.columns:
            column_letter = col[0].column_letter
            ws.column_dimensions[column_letter].width = desired_width

    # Identify rows where each employee's payslip starts
    employee_start_indices = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row[0] and re.match(r'\w+ \w+ \w+\s+\(\d+\)', str(row[0])):
            employee_start_indices.append(i + 1)  # adjust to 1-based indexing used by openpyxl

    # Ensure we cover the entire worksheet
    employee_start_indices.append(ws.max_row + 1)

    # Split and save each employee's data, preserving format
    saved_files = []
    for idx in range(len(employee_start_indices) - 1):
        start_idx = employee_start_indices[idx]
        end_idx = employee_start_indices[idx + 1] - 1

        # Create a new workbook and worksheet
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active

        # Copy rows to the new worksheet
        for row in ws.iter_rows(min_row=start_idx, max_row=end_idx):
            for cell in row:
                new_cell = new_ws.cell(row=cell.row - start_idx + 1, column=cell.column, value=cell.value)
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

        for i in range(1, ws.max_column+1):
            # преобразовываем индекс столбца в его букву
            letter = get_column_letter(i)
            # получаем ширину столбца
            ws.column_dimensions[letter].width = 0.5




        # Copy merged cells
        for merged_cell in ws.merged_cells.ranges:
            if merged_cell.min_row >= start_idx and merged_cell.max_row <= end_idx:
                new_ws.merge_cells(
                    start_row=merged_cell.min_row - start_idx + 1,
                    start_column=merged_cell.min_col,
                    end_row=merged_cell.max_row - start_idx + 1,
                    end_column=merged_cell.max_col
                )


        # Extract employee name and ID for the filename
        employee_info = ws.cell(row=start_idx, column=1).value
        employee_name_id = re.search(r'(\w+ \w+ \w+\s+\(\d+\))', employee_info).group(1)
        filename = f'{output_path}/{employee_name_id}.xlsx'





        # Save the new workbook
        new_wb.save(filename)
        normal_width(filename)

        saved_files.append(filename)


    return(saved_files)
    


# In[9]:


file_path = os.path.join(os.path.dirname(__file__), 'Расчетный лист.xlsx')

parent_id = upload_and_send_document(file_path, 'ipetrov')

output_path = os.path.join(os.path.dirname(__file__), 'result')

output_files = split_xlsx(file_path, output_path)

upload_and_send_document(output_files[0], 'izubov', parent_id)

print(output_files)





